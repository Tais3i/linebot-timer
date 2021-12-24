from json import load
from flask import Flask, request, abort

from openpyxl import load_workbook

from linebot import (
    LineBotApi, WebhookHandler
)
from linebot.exceptions import (
    InvalidSignatureError
)
from linebot.models import     (
    MessageEvent, TextMessage, TextSendMessage,
)

app = Flask(__name__)

line_bot_api = LineBotApi('T9giHmlCfNjcHfexEI+U58VpDeMfElD/aXdga7ZckRrzfSsBDL2I1XXHkrIfeLTuaE+CfuTw5GSUzzmF5tOovfBItI3GyjkchU4lfAjzAf2hc/a25GIHfXQWnGagGMKtg8jqsjuagRg9maCqEik97wdB04t89/1O/w1cDnyilFU=')
handler = WebhookHandler('6e937500d48e9a084c0e940419aa3358')

start_times = {}

@app.route("/")
def test():
    return "OK"

@app.route("/callback", methods=['POST'])
def callback():
    # get X-Line-Signature header value
    signature = request.headers['X-Line-Signature']

    # get request body as text
    body = request.get_data(as_text=True)
    app.logger.info("Request body: " + body)

    # handle webhook body
    try:
        handler.handle(body, signature)
    except InvalidSignatureError:
        print("Invalid signature. Please check your channel access token/channel secret.")
        abort(400)

    return 'OK'

import time
import datetime
wb = load_workbook(filename='chatbot.xlsx')
ws = wb['勉強時間']
d_today = datetime.date.today()

@handler.add(MessageEvent, message=TextMessage)
def handle_message(event):
    if event.message.text == "開始":
        i = 0
        i += 1
        ws[f'A{i}'] = d_today
        start_times[event.source.user_id] = time.time()
        return line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(text="計測を開始します"))
    if event.message.text == "終了":
        d = 0
        d += 1
        elapsed_time = int(time.time() - start_times[event.source.user_id])
        del start_times[event.source.user_id]
        hour = elapsed_time // 3600
        minute = (elapsed_time % 3600) // 60
        second = elapsed_time % 60
        ws[f'B{d}'] = "{hour}時間{minute}分{second}秒"
        return line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(text=f"ただいまの勉強時間は{hour}時間{minute}分{second}秒でした。"))



if __name__ == "__main__":
    app.run()